#!/usr/bin/env python3
"""Build the tabletop unit-activity master table (post-V2 punchlist item 0).

One row per unit across the FULL manifest (ADR 0005) — every reconstructed
phase's battle file, not just the July 3 afternoon file — with in-build
movement and casualty metrics computed from the keyframes, research-coverage
signals mined from docs/research and the V2 claims corpus, and empty
owner-consultation columns to be layered in.

Manifest-driven (not hardcoded): the phase list comes from
app/Assets/StreamingAssets/Atlas/battle-manifest.json, so a unit authored in
ANY reconstructed phase (Gamble/Devin/Calef in july1-morning, Harrow's
children and Ziegler's Grove in july3, etc.) gets a row — previously only
gettysburg-july3.json was read, so every day-expansion-authored unit was
invisible to the computed columns (decomposition-wave-1.md §10 pickup 2).
A unit appearing in multiple phases (a continuing brigade/battery) gets ONE
row: its keyframes/events are concatenated across phases in manifest day/
phase order, so "Start strength" is its earliest appearance's t=0 and "End
strength" is its latest appearance's final keyframe — the same movement-
metric math as before, run over the unit's full known arc instead of one
phase's slice. A new "Phases" column names every phase the unit appears in.

Per the unit-truth-spec scope ruling, the table also carries one row per
NOT-YET-CAST entry of the full three-day OOB register
(docs/reconstruction/audit/oob-register.json) — cast-status column on every
row; computed columns blank where no build data exists.

Regenerable: computed columns are overwritten on rebuild; consultation
columns (owner-edited) are PRESERVED by re-reading the existing workbook
before writing, keyed by unit id.

Usage:
    cd reconstruction && uv run --with openpyxl python scripts/build_unit_audit.py
Output:
    docs/reconstruction/audit/unit-master-table.xlsx (+ .csv companion
    of computed columns only, for git diffs)
"""

import csv
import json
import math
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
MANIFEST = ROOT / "app/Assets/StreamingAssets/Atlas/battle-manifest.json"
BATTLE_DIR = ROOT / "app/Assets/Battle"
CLAIMS = ROOT / "reconstruction/claims/angle.claims.json"
RESEARCH_DIRS = [ROOT / "docs/research"]
OUT_DIR = ROOT / "docs/reconstruction/audit"
XLSX = OUT_DIR / "unit-master-table.xlsx"
CSV_OUT = OUT_DIR / "unit-master-table.csv"
REGISTER = OUT_DIR / "oob-register.json"
# Dossier overlay: dossier passes commit consultation-layer values here,
# keyed by master-table Unit ID (= build id for in-build rows, register id
# otherwise). Overlay values win over values preserved from the existing
# workbook, so the consultation layer for audited units regenerates
# deterministically from a committed input instead of from workbook state.
OVERLAY = OUT_DIR / "dossier-overlay.json"

MOVE_THRESHOLD_M = 100.0  # path length below this ≈ dressing/jitter, not movement

CONSULT_COLS = [
    "Class (attested-static / unauthored / authored-ok)",
    "Historical actions summary (owner/consult)",
    "Historical casualties (killed/wounded/missing, source)",
    "Proposed changes (keyframes/events to author)",
    "Priority (1-3)",
    "Notes",
    # unit-truth-spec.md layers (owner-consultation, tier-prefixed facts)
    "EC1 identity/command",
    "EC2 engaged strength",
    "EC3 position anchors",
    "EC4 movement legs",
    "EC5 activity record",
    "EC6 casualty apportionment",
    "Target T-level",
    "Achieved T-level (audited)",
]


def arm_of(name: str) -> str:
    n = name.lower()
    if "battery" in n or "artillery" in n:
        return "artillery"
    if "cavalry" in n:
        return "cavalry"
    return "infantry"


def research_token(name: str) -> str:
    """A distinctive grep token: commander possessive stem or regiment number."""
    m = re.match(r"^(.*?)'s\b", name)
    if m:
        return m.group(1).split()[-1]  # "Garnett's Brigade" -> "Garnett"
    m = re.match(r"^(\d+(?:st|nd|rd|th)\s+\w+)", name)
    if m:
        return m.group(1)  # "13th Vermont Infantry" -> "13th Vermont"
    return name.split("(")[0].strip().split(",")[0]


def path_metrics(kfs):
    path = 0.0
    for a, b in zip(kfs, kfs[1:]):
        path += math.hypot(b["x"] - a["x"], b["z"] - a["z"])
    net = math.hypot(kfs[-1]["x"] - kfs[0]["x"], kfs[-1]["z"] - kfs[0]["z"])
    return round(path, 1), round(net, 1)


def reconstructed_phases(manifest_path: Path = MANIFEST) -> list[tuple[str, str]]:
    """(phaseId, battleFileName) for every reconstructed phase, in the
    manifest's own day/phase order (day.date is required strictly
    increasing; phases within a day are authored in clock order) —
    manifest-driven, so a newly-added phase file is picked up with no code
    change here."""
    manifest = json.loads(manifest_path.read_text())
    out = []
    for day in manifest["days"]:
        for phase in day["phases"]:
            if phase["status"] == "reconstructed":
                out.append((phase["id"], phase["battle"]))
    return out


def load_phase_battles(
    manifest_path: Path = MANIFEST, battle_dir: Path = BATTLE_DIR
) -> list[tuple[str, dict]]:
    """(phaseId, battleFileJson) for every reconstructed phase, in manifest
    order."""
    return [
        (phase_id, json.loads((battle_dir / battle_file).read_text()))
        for phase_id, battle_file in reconstructed_phases(manifest_path)
    ]


def merge_units_across_phases(
    phase_battles: list[tuple[str, dict]],
) -> tuple[dict[str, dict], dict[str, list], set[str]]:
    """Merge every phase's units into one row per unit id.

    A unit appearing in multiple phases (a continuing brigade/battery) is
    concatenated: keyframes and events accumulate across phases in manifest
    order, so all the existing per-unit movement/casualty math (path
    length, start/end strength, confidence tallies) runs over the unit's
    full known arc. Name/side/parent/echelon-defining fields come from the
    LAST phase the unit appears in (the most current authored state — e.g.
    a brigade renamed "... minus the Nth ___" by a later decomposition
    wave). Returns (merged_units, merged_events_by_unit, all_parent_ids).
    """
    merged: dict[str, dict] = {}
    events_by_unit: dict[str, list] = {}
    parents: set[str] = set()
    for phase_id, battle in phase_battles:
        for u in battle["units"]:
            if u.get("parent"):
                parents.add(u["parent"])
            row = merged.setdefault(
                u["id"],
                {"id": u["id"], "keyframes": [], "phases": []},
            )
            # last-phase-wins for identity fields; accumulate keyframes/phases
            row["name"] = u["name"]
            row["side"] = u["side"]
            row["parent"] = u.get("parent")
            row["keyframes"].extend(u["keyframes"])
            row["phases"].append(phase_id)
        for e in battle["events"]:
            uid = e.get("unitId")
            if uid:
                events_by_unit.setdefault(uid, []).append(e)
    return merged, events_by_unit, parents


def main():
    phase_battles = load_phase_battles()
    if not phase_battles:
        raise SystemExit(f"no reconstructed phases found in {MANIFEST}")
    units_by_id, events_by_unit, parents = merge_units_across_phases(phase_battles)
    units = list(units_by_id.values())
    claims = json.loads(CLAIMS.read_text())
    claim_subjects = {}
    for c in claims["claims"] if isinstance(claims, dict) else claims:
        claim_subjects.setdefault(c["subjectId"], 0)
        claim_subjects[c["subjectId"]] += 1

    research_texts = {}
    for d in RESEARCH_DIRS:
        for f in sorted(d.glob("*.md")):
            research_texts[f.name] = f.read_text(errors="ignore")

    rows = []
    for u in units:
        kfs = u["keyframes"]
        path, net = path_metrics(kfs)
        moves = path >= MOVE_THRESHOLD_M
        conf = [k.get("confidence", "") for k in kfs]
        documented = sum(1 for c in conf if c == "documented")
        inferred = sum(1 for c in conf if c == "inferred")
        unknown = sum(1 for c in conf if c == "unknown")
        cited = sum(1 for k in kfs if k.get("citation"))
        s0, s1 = kfs[0]["strength"], kfs[-1]["strength"]
        unit_events = events_by_unit.get(u["id"], [])
        token = research_token(u["name"])
        mention_docs = [
            fn for fn, txt in research_texts.items() if token and token in txt
        ]
        echelon = (
            "child-regiment"
            if u.get("parent")
            else ("parent-brigade" if u["id"] in parents else "standalone")
        )
        formations = "/".join(sorted({k["formation"] for k in kfs}))
        has_info = bool(mention_docs) or documented > 0 or u["id"] in claim_subjects
        # computed T-level floor from build data alone (audit raises it):
        # T5 = Angle cast (has V2 claims); T1 = documented endpoint anchors;
        # T0 = everything else. T2-T4 are only reachable via the audit.
        if u["id"] in claim_subjects:
            t_computed = "T5"
        elif kfs and conf and conf[0] == "documented" and conf[-1] == "documented":
            t_computed = "T1"
        else:
            t_computed = "T0"
        # suggested target per spec: >=T2 all, >=T3 movers, >=T4 engaged
        engaged = len(unit_events) > 0 or (s0 - s1) > 0
        t_target = "T4" if engaged else ("T3" if moves else "T2")
        rows.append(
            {
                "Unit ID": u["id"],
                "Name": u["name"],
                "Side": u["side"],
                "Echelon": echelon,
                "Arm": arm_of(u["name"]),
                "Phases": "/".join(u["phases"]),
                "Start strength": s0,
                "End strength": s1,
                "In-build casualties (n)": s0 - s1,
                "Keyframes": len(kfs),
                "KF documented": documented,
                "KF inferred": inferred,
                "KF unknown": unknown,
                "KF cited": cited,
                "Path length (m)": path,
                "Net displacement (m)": net,
                "Moves in build?": "yes" if moves else "no",
                "Formations seen": formations,
                "Fire events": len(unit_events),
                "V2 claims": claim_subjects.get(u["id"], 0),
                "Research docs mentioning": len(mention_docs),
                "Research token": token,
                "Movement info available?": "yes" if has_info else "VERIFY",
                "T-level (computed floor)": t_computed,
                "T-level target (suggested)": t_target,
            }
        )

    # --- full-OOB register overlay (scope ruling: full-battle grain) ---
    register = json.loads(REGISTER.read_text())["entries"] if REGISTER.exists() else []
    cast_to_register = {e["castStatus"]: e["id"] for e in register
                        if e["castStatus"] != "not-yet-cast"}
    for r in rows:
        r["Cast status"] = "in-build"
        r["Register entry"] = cast_to_register.get(r["Unit ID"], "")

    for e in register:
        if e["castStatus"] != "not-yet-cast":
            continue
        token = research_token(e["name"])
        mention_docs = [fn for fn, txt in research_texts.items()
                        if token and token in txt]
        rows.append({
            "Unit ID": e["id"],
            "Name": e["name"],
            "Side": e["side"],
            "Echelon": e["echelon"],
            "Arm": e["arm"],
            "Phases": "",
            "Start strength": "",
            "End strength": "",
            "In-build casualties (n)": "",
            "Keyframes": "",
            "KF documented": "",
            "KF inferred": "",
            "KF unknown": "",
            "KF cited": "",
            "Path length (m)": "",
            "Net displacement (m)": "",
            "Moves in build?": "",
            "Formations seen": "",
            "Fire events": "",
            "V2 claims": "",
            "Research docs mentioning": len(mention_docs),
            "Research token": token,
            "Movement info available?": "yes" if mention_docs else "VERIFY",
            "T-level (computed floor)": "",
            "T-level target (suggested)": "T2",
            "Cast status": "not-yet-cast",
            "Register entry": e["id"],
        })

    rows.sort(key=lambda r: (r["Side"], r["Arm"], r["Echelon"], r["Name"]))

    # preserve owner consultation columns from an existing workbook
    preserved = {}
    if XLSX.exists():
        from openpyxl import load_workbook

        wb_old = load_workbook(XLSX, data_only=True)
        ws_old = wb_old["Units"]
        headers = [c.value for c in ws_old[1]]
        for r in ws_old.iter_rows(min_row=2, values_only=True):
            rec = dict(zip(headers, r))
            uid = rec.get("Unit ID")
            if uid:
                preserved[uid] = {c: rec.get(c) for c in CONSULT_COLS if c in rec}

    # dossier overlay wins over workbook-preserved values (see OVERLAY note)
    if OVERLAY.exists():
        overlay = json.loads(OVERLAY.read_text())
        for uid, cols in overlay.get("units", {}).items():
            unknown = set(cols) - set(CONSULT_COLS)
            if unknown:
                raise SystemExit(
                    f"dossier-overlay.json: unknown consultation column(s) "
                    f"{sorted(unknown)} on unit {uid}"
                )
            preserved.setdefault(uid, {}).update(cols)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Units"
    computed_cols = list(rows[0].keys())
    # "Phases" (new, manifest coverage) sits before Start/End strength, so the
    # casualty-% formula insertion point shifts from 7 to 8 columns in.
    STRENGTH_COLS = 8  # Unit ID..Phases..Start strength..End strength
    all_cols = (
        computed_cols[:STRENGTH_COLS]
        + ["In-build casualty %"]
        + computed_cols[STRENGTH_COLS:]
        + CONSULT_COLS
    )
    arial = Font(name="Arial", size=10)
    bold = Font(name="Arial", size=10, bold=True)
    yellow = PatternFill("solid", fgColor="FFF2CC")
    red_font = Font(name="Arial", size=10, color="CC0000", bold=True)

    ws.append(all_cols)
    for c in ws[1]:
        c.font = bold

    # column letters for the live casualty-% formula, computed from all_cols
    # so an inserted/reordered column (e.g. "Phases") never desyncs the
    # formula from "Start strength"/"End strength" by magic-number drift.
    start_col = get_column_letter(all_cols.index("Start strength") + 1)
    end_col = get_column_letter(all_cols.index("End strength") + 1)
    pct_col_idx = all_cols.index("In-build casualty %") + 1

    for i, r in enumerate(rows, start=2):
        vals = [r[c] for c in computed_cols[:STRENGTH_COLS]]
        vals.append(
            f"=IF({start_col}{i}=0,0,1-{end_col}{i}/{start_col}{i})"
        )  # casualty % as a live formula
        vals += [r[c] for c in computed_cols[STRENGTH_COLS:]]
        p = preserved.get(r["Unit ID"], {})
        vals += [p.get(c) for c in CONSULT_COLS]
        ws.append(vals)
        for c in ws[i]:
            c.font = arial
        ws.cell(row=i, column=pct_col_idx).number_format = "0.0%"
        if r["Movement info available?"] == "VERIFY":
            ws.cell(row=i, column=all_cols.index("Movement info available?") + 1).font = red_font
        for cc in CONSULT_COLS:
            ws.cell(row=i, column=all_cols.index(cc) + 1).fill = yellow

    # widths keyed by column NAME (not position) so inserting/reordering a
    # column never silently mis-widths an unrelated one.
    named_widths = {
        "Unit ID": 22, "Name": 38, "Side": 12, "Echelon": 15, "Arm": 10,
        "Net displacement (m)": 22, "V2 claims": 16,
    }
    for idx, col in enumerate(all_cols, start=1):
        width = named_widths.get(col, 26 if col in CONSULT_COLS else 13)
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}{len(rows)+1}"

    legend = wb.create_sheet("Legend")
    legend_rows = [
        ["Tabletop unit-activity audit — master table (punchlist item 0)"],
        [""],
        ["Computed columns (white) are regenerated by "
         "reconstruction/scripts/build_unit_audit.py — do not hand-edit."],
        ["Yellow columns are the owner-consultation layer and SURVIVE "
         "regeneration (matched by Unit ID)."],
        ["'Moves in build?' = path length >= "
         f"{MOVE_THRESHOLD_M:.0f} m across all keyframes."],
        ["'Movement info available?' = any of: documented keyframes, V2 "
         "claims about the unit, or research-doc mentions of the token. "
         "VERIFY (red) = none found — per the owner, these units are "
         "likely researched and the gap is ours; verify before assuming "
         "the record is silent."],
        ["'In-build casualty %' is a live formula on start/end strength "
         "(the strengths themselves come from authored keyframes)."],
        ["Class values: attested-static (record supports stillness) / "
         "unauthored (recorded movement missing from the build) / "
         "authored-ok (current track already right)."],
        ["Example consultation row: see csa-garnett (filled as a format "
         "example from the V2 corpus — ED-2/ED-8, casualty profiles)."],
    ]
    for lr in legend_rows:
        legend.append(lr)
    for row in legend.iter_rows():
        for c in row:
            c.font = arial
    legend.column_dimensions["A"].width = 110

    # one example consultation row, from data already in the repo
    if "csa-garnett" not in preserved or not any(
        (preserved.get("csa-garnett") or {}).values()
    ):
        for i, r in enumerate(rows, start=2):
            if r["Unit ID"] == "csa-garnett":
                ex = [
                    "authored-ok",
                    "Advanced in Pickett's front line from Spangler's Woods; "
                    "halted under fire at the Emmitsburg Road fences "
                    "(~15:16-15:18, ED-12/claim-fences-high-climb); crossed "
                    "under canister; reached the wall; repulsed ~15:25+ "
                    "(V2 corpus: 7 segments, ED-2 timing frame).",
                    "In-build decay 1480->539 over the slice; V2 casualty "
                    "profiles reconcile exactly (e.g. "
                    "cas-garnett-angle-approach: 510, rising, "
                    "musketry/canister). Battle-total K/W/M per OOB research "
                    "doc - layer in.",
                    "None needed for the slice window (Angle cast).",
                    "3",
                    "Format example row - overwrite freely.",
                ]
                for j, cc in enumerate(CONSULT_COLS):
                    ws.cell(row=i, column=all_cols.index(cc) + 1, value=ex[j])
                break

    wb.save(XLSX)

    with open(CSV_OUT, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=computed_cols)
        w.writeheader()
        w.writerows(rows)

    in_build = sum(1 for r in rows if r["Cast status"] == "in-build")
    not_cast = len(rows) - in_build
    phase_ids = [p for p, _ in reconstructed_phases()]
    print(f"phases read (manifest-driven): {len(phase_ids)}: {', '.join(phase_ids)}")
    print(f"rows: {len(rows)} | in-build: {in_build} | not-yet-cast (register): {not_cast}")
    movers = sum(1 for r in rows if r["Moves in build?"] == "yes")
    verify = [r["Unit ID"] for r in rows if r["Movement info available?"] == "VERIFY"]
    print(f"units: {len(rows)} | move in build: {movers} | static: {len(rows)-movers}")
    print(f"VERIFY (no coverage signal found): {len(verify)}")
    for v in verify[:15]:
        print("  ", v)
    print(f"wrote {XLSX.relative_to(ROOT)} and {CSV_OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
